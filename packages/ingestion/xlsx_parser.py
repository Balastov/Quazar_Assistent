import io
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell


def _cell_value(cell: Cell) -> str:
    value = cell.value
    if value is None:
        return ""
    if isinstance(value, bool):
        return "да" if value else "нет"
    return str(value).strip()


def extract_xlsx(content: bytes, filename: str) -> str:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    parts: list[str] = [f"# Файл: {filename}"]

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parts.append(f"\n## Лист: {sheet_name}")

        rows: list[list[str]] = []
        for row in ws.iter_rows():
            cells = [_cell_value(c) for c in row]
            if any(cells):
                rows.append(cells)

        if not rows:
            parts.append("(пустой лист)")
            continue

        max_cols = max(len(r) for r in rows)
        for row in rows:
            while len(row) < max_cols:
                row.append("")

        header = rows[0]
        parts.append("| " + " | ".join(header) + " |")
        parts.append("| " + " | ".join(["---"] * max_cols) + " |")
        for row in rows[1:]:
            parts.append("| " + " | ".join(row) + " |")

    wb.close()
    return "\n".join(parts)
